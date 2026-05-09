"""Premium Excel exporter: multi-sheet, conditional formatting, charts, hyperlinks."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import Competitor, ScoredKeyword


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="555555")
TABLE_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DEFAULT_FONT = Font(name="Calibri", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

THIN_SIDE = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

SCORE_COLOR_SCALE = ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B",
)
OVERLAP_COLOR_SCALE = ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B",
)

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="F8CBAD")

BEST_COLUMNS = [
    ("Keyword", "name"),
    ("Score", "score"),
    ("Popularity", "popularity"),
    ("Competitor count", "competitor_count"),
    ("Avg competitor rank", "avg_competitor_ranking"),
    ("Competitors", "_competitors_joined"),
]

ADVANCED_COLUMNS = [
    ("Keyword", "name"),
    ("Composite Score", "composite_score"),
    ("Popularity", "popularity"),
    ("Difficulty", "difficulty"),
    ("Opportunity", "opportunity"),
    ("Competitor Count", "competitor_count"),
    ("Length Class", "length_class"),
    ("Is Brand", "is_brand"),
    ("Is Local", "is_local"),
    ("Competitors", "_competitors_joined"),
]

ADVANCED_TIER_ORDER = ["WINNERS", "EASY_WINS", "HIGH_VALUE", "HIDDEN_GEMS", "LOCAL_PLAYS"]


def to_excel_pro(
    classified_simple: dict[str, list[ScoredKeyword]],
    competitors: list[Competitor],
    output_path: str | Path,
    classified_advanced: dict[str, list] | None = None,
    title: str = "ASO Niche Report",
    subtitle: str = "",
    country_code: int = 25,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    _build_summary_sheet(wb, classified_simple, title, subtitle)

    for tier_name in ("BEST", "MEDIUM"):
        items = classified_simple.get(tier_name, [])
        _build_tier_sheet(wb, tier_name, items)

    if classified_advanced:
        for tier_name in ADVANCED_TIER_ORDER:
            items = classified_advanced.get(tier_name)
            if not items:
                continue
            _build_advanced_sheet(wb, tier_name, items)

    _build_competitors_sheet(wb, competitors, country_code)
    _build_charts_sheet(wb, classified_simple)

    wb.save(output_path)
    return output_path


def _build_summary_sheet(
    wb: Workbook,
    classified_simple: dict[str, list[ScoredKeyword]],
    title: str,
    subtitle: str,
) -> None:
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    if subtitle:
        ws.merge_cells("A2:E2")
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

    tier_table_start_row = 4
    ws.cell(row=tier_table_start_row, column=1, value="Tier").font = TABLE_HEADER_FONT
    ws.cell(row=tier_table_start_row, column=2, value="Count").font = TABLE_HEADER_FONT
    ws.cell(row=tier_table_start_row, column=1).fill = HEADER_FILL
    ws.cell(row=tier_table_start_row, column=2).fill = HEADER_FILL
    ws.cell(row=tier_table_start_row, column=1).alignment = CENTER
    ws.cell(row=tier_table_start_row, column=2).alignment = CENTER

    ordered_tiers = ["BEST", "MEDIUM", "TRASH"]
    extra_tiers = [t for t in classified_simple.keys() if t not in ordered_tiers]
    for tier_name in ordered_tiers + extra_tiers:
        if tier_name not in classified_simple:
            continue
        row_idx = tier_table_start_row + 1 + (ordered_tiers + extra_tiers).index(tier_name)
        ws.cell(row=row_idx, column=1, value=tier_name).alignment = LEFT
        ws.cell(row=row_idx, column=2, value=len(classified_simple[tier_name])).alignment = CENTER
        ws.cell(row=row_idx, column=1).border = CELL_BORDER
        ws.cell(row=row_idx, column=2).border = CELL_BORDER

    top_table_start_row = tier_table_start_row + len(ordered_tiers + extra_tiers) + 3
    ws.cell(row=top_table_start_row - 1, column=1, value="Top 10 Keywords by Score").font = Font(
        name="Calibri", size=13, bold=True, color="1F4E78"
    )

    headers = ["Keyword", "Tier", "Score", "Popularity", "Competitors"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=top_table_start_row, column=col_idx, value=h)
        c.font = TABLE_HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    all_kws: list[ScoredKeyword] = []
    for tier_items in classified_simple.values():
        all_kws.extend(tier_items)
    all_kws_sorted = sorted(all_kws, key=lambda k: k.score, reverse=True)[:10]

    for i, kw in enumerate(all_kws_sorted, start=1):
        r = top_table_start_row + i
        ws.cell(row=r, column=1, value=kw.name).alignment = LEFT
        ws.cell(row=r, column=2, value=kw.tier).alignment = CENTER
        ws.cell(row=r, column=3, value=kw.score).alignment = CENTER
        ws.cell(row=r, column=4, value=kw.popularity).alignment = CENTER
        ws.cell(row=r, column=5, value=kw.competitor_count).alignment = CENTER
        for col_idx in range(1, 6):
            ws.cell(row=r, column=col_idx).border = CELL_BORDER

    if all_kws_sorted:
        score_col_letter = get_column_letter(3)
        score_range = f"{score_col_letter}{top_table_start_row + 1}:{score_col_letter}{top_table_start_row + len(all_kws_sorted)}"
        ws.conditional_formatting.add(score_range, SCORE_COLOR_SCALE)

    _autofit_columns(ws, max_width=60)


def _build_tier_sheet(wb: Workbook, tier_name: str, items: list[ScoredKeyword]) -> None:
    ws = wb.create_sheet(tier_name)
    headers = [h for h, _ in BEST_COLUMNS]
    _write_header_row(ws, headers)

    for i, kw in enumerate(items, start=2):
        row_values = _scored_keyword_row(kw)
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = LEFT if col_idx in (1, len(headers)) else CENTER

    ws.freeze_panes = "A2"

    if items:
        last_row = len(items) + 1
        score_col_letter = get_column_letter(2)
        rank_col_letter = get_column_letter(5)

        score_range = f"{score_col_letter}2:{score_col_letter}{last_row}"
        rank_range = f"{rank_col_letter}2:{rank_col_letter}{last_row}"

        ws.conditional_formatting.add(score_range, SCORE_COLOR_SCALE)

        ws.conditional_formatting.add(
            rank_range,
            CellIsRule(operator="lessThanOrEqual", formula=["10"], fill=GREEN_FILL),
        )
        ws.conditional_formatting.add(
            rank_range,
            CellIsRule(operator="between", formula=["11", "30"], fill=YELLOW_FILL),
        )
        ws.conditional_formatting.add(
            rank_range,
            CellIsRule(operator="greaterThan", formula=["30"], fill=RED_FILL),
        )

    _autofit_columns(ws, max_width=60)


def _build_advanced_sheet(wb: Workbook, tier_name: str, items: list[Any]) -> None:
    ws = wb.create_sheet(tier_name)
    headers = [h for h, _ in ADVANCED_COLUMNS]
    _write_header_row(ws, headers)

    for i, kw in enumerate(items, start=2):
        row_values = _advanced_keyword_row(kw)
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.alignment = LEFT if col_idx in (1, len(headers)) else CENTER

    ws.freeze_panes = "A2"

    if items:
        last_row = len(items) + 1
        score_col_letter = get_column_letter(2)
        score_range = f"{score_col_letter}2:{score_col_letter}{last_row}"
        ws.conditional_formatting.add(score_range, SCORE_COLOR_SCALE)

    _autofit_columns(ws, max_width=60)


def _build_competitors_sheet(
    wb: Workbook,
    competitors: list[Competitor],
    country_code: int,
) -> None:
    ws = wb.create_sheet("COMPETITORS")
    headers = [
        "App ID",
        "Name",
        "Category",
        "Seed overlap %",
        "Matched seeds",
        "Validated",
        "Keywords scraped",
    ]
    _write_header_row(ws, headers)

    link_font = Font(name="Calibri", size=11, color="1F4E78", underline="single")

    for i, c in enumerate(competitors, start=2):
        url = f"https://www.upup.com/app/{c.app_id}/ios-aso?market=1&country={country_code}"
        app_id_cell = ws.cell(row=i, column=1, value=c.app_id)
        app_id_cell.hyperlink = url
        app_id_cell.font = link_font
        app_id_cell.alignment = LEFT

        ws.cell(row=i, column=2, value=c.name).alignment = LEFT
        ws.cell(row=i, column=3, value=c.category).alignment = LEFT
        ws.cell(row=i, column=4, value=round(c.seed_overlap * 100, 1)).alignment = CENTER
        ws.cell(row=i, column=5, value=", ".join(c.matched_seeds)).alignment = LEFT
        ws.cell(row=i, column=6, value="YES" if c.validated else "NO").alignment = CENTER
        ws.cell(row=i, column=7, value=len(c.keywords)).alignment = CENTER

    ws.freeze_panes = "A2"

    if competitors:
        last_row = len(competitors) + 1
        overlap_col_letter = get_column_letter(4)
        overlap_range = f"{overlap_col_letter}2:{overlap_col_letter}{last_row}"
        ws.conditional_formatting.add(overlap_range, OVERLAP_COLOR_SCALE)

    _autofit_columns(ws, max_width=60)


def _build_charts_sheet(
    wb: Workbook,
    classified_simple: dict[str, list[ScoredKeyword]],
) -> None:
    ws = wb.create_sheet("Charts")

    ws["A1"] = "Visual Overview"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    ws["A3"] = "Top 15 BEST Keywords"
    ws["A3"].font = Font(name="Calibri", size=12, bold=True, color="1F4E78")

    best_items = sorted(
        classified_simple.get("BEST", []),
        key=lambda k: k.score,
        reverse=True,
    )[:15]

    ws.cell(row=4, column=1, value="Keyword").font = TABLE_HEADER_FONT
    ws.cell(row=4, column=2, value="Score").font = TABLE_HEADER_FONT
    ws.cell(row=4, column=1).fill = HEADER_FILL
    ws.cell(row=4, column=2).fill = HEADER_FILL

    for i, kw in enumerate(best_items, start=5):
        ws.cell(row=i, column=1, value=kw.name)
        ws.cell(row=i, column=2, value=kw.score)

    if best_items:
        bar = BarChart()
        bar.type = "bar"
        bar.style = 11
        bar.title = "Top BEST Keywords by Score"
        bar.y_axis.title = "Keyword"
        bar.x_axis.title = "Score"
        bar.height = 12
        bar.width = 22

        data_ref = Reference(
            ws,
            min_col=2,
            min_row=4,
            max_row=4 + len(best_items),
            max_col=2,
        )
        cat_ref = Reference(
            ws,
            min_col=1,
            min_row=5,
            max_row=4 + len(best_items),
        )
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cat_ref)
        ws.add_chart(bar, "D3")

    pie_table_start_row = max(4 + len(best_items) + 3, 22)
    ws.cell(row=pie_table_start_row - 1, column=1, value="Tier Distribution").font = Font(
        name="Calibri", size=12, bold=True, color="1F4E78"
    )

    ws.cell(row=pie_table_start_row, column=1, value="Tier").font = TABLE_HEADER_FONT
    ws.cell(row=pie_table_start_row, column=2, value="Count").font = TABLE_HEADER_FONT
    ws.cell(row=pie_table_start_row, column=1).fill = HEADER_FILL
    ws.cell(row=pie_table_start_row, column=2).fill = HEADER_FILL

    tier_rows = []
    for tier_name in ("BEST", "MEDIUM", "TRASH"):
        if tier_name in classified_simple:
            tier_rows.append((tier_name, len(classified_simple[tier_name])))

    for i, (tier_name, count) in enumerate(tier_rows, start=1):
        ws.cell(row=pie_table_start_row + i, column=1, value=tier_name)
        ws.cell(row=pie_table_start_row + i, column=2, value=count)

    if tier_rows:
        pie = PieChart()
        pie.title = "Keywords by Tier"
        pie.height = 10
        pie.width = 14

        data_ref = Reference(
            ws,
            min_col=2,
            min_row=pie_table_start_row,
            max_row=pie_table_start_row + len(tier_rows),
            max_col=2,
        )
        labels_ref = Reference(
            ws,
            min_col=1,
            min_row=pie_table_start_row + 1,
            max_row=pie_table_start_row + len(tier_rows),
        )
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels_ref)
        pie.dataLabels = DataLabelList(showPercent=True)
        ws.add_chart(pie, f"D{pie_table_start_row}")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12


def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22


def _scored_keyword_row(kw: ScoredKeyword) -> list[Any]:
    return [
        kw.name,
        kw.score,
        kw.popularity,
        kw.competitor_count,
        kw.avg_competitor_ranking,
        ", ".join(kw.competitors),
    ]


def _advanced_keyword_row(kw: Any) -> list[Any]:
    def g(name: str, default: Any = "") -> Any:
        if isinstance(kw, dict):
            return kw.get(name, default)
        return getattr(kw, name, default)

    competitors_raw = g("competitors", [])
    if isinstance(competitors_raw, (list, tuple, set)):
        competitors_joined = ", ".join(str(c) for c in competitors_raw)
    else:
        competitors_joined = str(competitors_raw) if competitors_raw else ""

    return [
        g("name", ""),
        g("composite_score", g("score", 0)),
        g("popularity", 0),
        g("difficulty", 0),
        g("opportunity", 0),
        g("competitor_count", 0),
        g("length_class", ""),
        "YES" if g("is_brand", False) else "NO",
        "YES" if g("is_local", False) else "NO",
        competitors_joined,
    ]


def _autofit_columns(ws: Worksheet, max_width: int = 60) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.column is None:
                continue
            value = cell.value
            if value is None:
                continue
            text = str(value)
            longest_line = max((len(line) for line in text.splitlines()), default=len(text))
            current = widths.get(cell.column, 0)
            if longest_line > current:
                widths[cell.column] = longest_line

    for col_idx, width in widths.items():
        letter = get_column_letter(col_idx)
        adjusted = min(max(width + 2, 10), max_width)
        ws.column_dimensions[letter].width = adjusted
