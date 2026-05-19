"""Export niche finder results to a styled .xlsx file."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── shared style constants ──────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT    = Font(name="Calibri", size=14, bold=True, color="1F3864")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ZEBRA_FILL    = PatternFill("solid", fgColor="F5F8FF")

TIER_STYLES = {
    "EXPLOSIVE": (PatternFill("solid", fgColor="FF4C4C"), Font(bold=True, color="FFFFFF")),
    "STRONG":    (PatternFill("solid", fgColor="FFD966"), Font(bold=True, color="7A4600")),
    "RISING":    (PatternFill("solid", fgColor="C6EFCE"), Font(bold=True, color="276221")),
    "weak":      (PatternFill("solid", fgColor="EFEFEF"), Font(color="888888")),
}

UPUP_APP_URL = "https://www.upup.com/app/{app_id}/ios-aso?market=1&country={country_id}"
LINK_FONT    = Font(color="0563C1", underline="single")

COLUMN_WIDTHS = {
    "Tier":         11,
    "Niche Score":  12,
    "Rank":          7,
    "Rank Jump":     11,
    "App Name":      32,
    "Developer":     28,
    "Genre":         18,
    "App Age (days)": 15,
    "Reviews":       12,
    "Rating":         8,
    "Price":          8,
    "New App":        9,
    "Organic":       10,
    "Release Date":  14,
    "upup Link":     14,
}


def _niches_to_df(niches: list[dict]) -> pd.DataFrame:
    rows = []
    for n in niches:
        rows.append({
            "Tier":           n["tier"],
            "Niche Score":    n["niche_score"],
            "Rank":           n["rank"],
            "Rank Jump":      n["rank_incr"],
            "App Name":       n["name"],
            "Developer":      n["developer"],
            "Genre":          n["genre"],
            "App Age (days)": n["app_age_days"],
            "Reviews":        n["rating_count"],
            "Rating":         n["rating"],
            "Price":          f"${n['price']:.2f}" if n["price"] else "Free",
            "New App":        "YES" if n["app_age_days"] <= 60 else "no",
            "Organic":        "YES" if not n["is_ad"] and not n["is_featured"] else "no",
            "Release Date":   n["release_date"],
            "upup Link":      UPUP_APP_URL.format(app_id=n["app_id"], country_id=24),
        })
    return pd.DataFrame(rows)


def _format_sheet(sheet, df: pd.DataFrame, niches: list[dict]) -> None:
    if df.empty:
        return

    headers = list(df.columns)
    last_row = sheet.max_row

    # header row
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # column widths
    for col_idx, name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(name, 14)

    tier_col       = headers.index("Tier") + 1
    score_col      = headers.index("Niche Score") + 1
    rank_jump_col  = headers.index("Rank Jump") + 1
    reviews_col    = headers.index("Reviews") + 1
    rating_col     = headers.index("Rating") + 1
    link_col       = headers.index("upup Link") + 1
    app_name_col   = headers.index("App Name") + 1

    numeric_cols = {score_col, rank_jump_col, reviews_col, rating_col,
                    headers.index("Rank") + 1, headers.index("App Age (days)") + 1}

    for row_idx in range(2, last_row + 1):
        zebra = (row_idx % 2 == 0)
        niche = niches[row_idx - 2] if (row_idx - 2) < len(niches) else {}

        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            if col_idx in numeric_cols or col_idx in (tier_col,):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = Alignment(vertical="center")

            if col_idx == score_col and isinstance(cell.value, (int, float)):
                cell.number_format = "0.0"
            if col_idx == reviews_col and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

            if zebra and col_idx != tier_col:
                cell.fill = ZEBRA_FILL

        # tier colour
        tier_val  = niche.get("tier", "weak")
        fill, fnt = TIER_STYLES.get(tier_val, TIER_STYLES["weak"])
        sheet.cell(row=row_idx, column=tier_col).fill = fill
        sheet.cell(row=row_idx, column=tier_col).font = fnt

        # rank jump colour
        incr = niche.get("rank_incr", 0)
        jcell = sheet.cell(row=row_idx, column=rank_jump_col)
        if incr >= 100:
            jcell.font = Font(bold=True, color="C00000")
        elif incr >= 30:
            jcell.font = Font(bold=True, color="E07300")
        else:
            jcell.font = Font(bold=True, color="276221")

        # upup link
        link_cell = sheet.cell(row=row_idx, column=link_col)
        if niche.get("app_id"):
            url = UPUP_APP_URL.format(app_id=niche["app_id"], country_id=24)
            link_cell.hyperlink = url
            link_cell.value = "Open ↗"
            link_cell.font = LINK_FONT
            link_cell.alignment = CENTER_ALIGN

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def export_niches_to_excel(niches: list[dict], output_path: str | Path) -> Path:
    """Write niche results to a formatted xlsx file. Returns the output path."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = _niches_to_df(niches)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="NICHES", index=False)
        _format_sheet(writer.sheets["NICHES"], df, niches)

    return output_path
