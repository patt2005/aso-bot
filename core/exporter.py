"""Export scored keywords + competitor list into a single .xlsx file.

Single KEYWORDS sheet (all tiers merged), with conditional formatting on each
numeric column so the user can scan green=good / red=bad at a glance.

The COMPETITORS sheet has clickable hyperlinks on each App ID.
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from core.models import Competitor, ScoredKeyword


UPUP_APP_URL = "https://www.upup.com/app/{app_id}/ios-aso?market=1&country={country}"
LINK_FONT = Font(color="0563C1", underline="single")
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

HEADER_FILL = PatternFill("solid", fgColor="374151")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="111827")

ZEBRA_FILL = PatternFill("solid", fgColor="F2F2F2")
TIER_BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
TIER_BEST_FONT = Font(bold=True, color="006100")
TIER_MED_FILL = PatternFill("solid", fgColor="FFEB9C")
TIER_MED_FONT = Font(bold=True, color="9C5700")
TIER_LOW_FILL = PatternFill("solid", fgColor="FFC7CE")
TIER_LOW_FONT = Font(color="9C0006")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

GREEN = "63BE7B"
YELLOW = "FFEB84"
RED = "F8696B"

FONT_GREEN = Font(name="Calibri", size=11, bold=True, color="0F7B0F")
FONT_YELLOW = Font(name="Calibri", size=11, bold=True, color="9C6500")
FONT_RED = Font(name="Calibri", size=11, bold=True, color="C00000")
FONT_ORANGE = Font(name="Calibri", size=11, bold=True, color="E07300")
FONT_NEUTRAL = Font(name="Calibri", size=11)


def _color_by_thresholds(value, thresholds, reverse=False):
    """Pick a font based on value vs thresholds.
    thresholds = (good_min, ok_min) when reverse=False -> higher is better
                 (good_max, ok_max) when reverse=True  -> lower is better
    Returns one of FONT_GREEN / FONT_YELLOW / FONT_RED.
    """
    if value is None:
        return FONT_NEUTRAL
    g, y = thresholds
    if not reverse:
        if value >= g:
            return FONT_GREEN
        if value >= y:
            return FONT_YELLOW
        return FONT_RED
    if value <= g:
        return FONT_GREEN
    if value <= y:
        return FONT_YELLOW
    return FONT_RED


def to_excel(
    classified: dict[str, list[ScoredKeyword]],
    competitors: list[Competitor],
    output_path: str | Path,
    country: int = 25,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    name_by_id = {c.app_id: (c.name or c.app_id) for c in competitors}

    all_items = (
        [(k, "BEST") for k in classified.get("BEST", [])]
        + [(k, "MEDIUM") for k in classified.get("MEDIUM", [])]
        + [(k, "LOW") for k in classified.get("TRASH", [])]
    )
    df_keywords = _scored_to_df(all_items, name_by_id)
    df_competitors = _competitors_to_df(competitors, country)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_keywords.to_excel(writer, sheet_name="KEYWORDS", index=False)
        df_competitors.to_excel(writer, sheet_name="COMPETITORS", index=False)

        _format_keyword_sheet(writer.sheets["KEYWORDS"], df_keywords)
        _hyperlink_competitor_ids(writer.sheets["COMPETITORS"], competitors, country)

    return output_path


def _scored_to_df(items: list[tuple[ScoredKeyword, str]], name_by_id: dict[str, str]) -> pd.DataFrame:
    rows = []
    for k, tier in items:
        comp_lines = []
        if k.top_search_apps:
            # Use top 5 apps from upup keyword search (actual ranking on this keyword)
            for app_name, rank_pos in k.top_search_apps[:5]:
                comp_lines.append(f"• {app_name} (#{rank_pos})")
        else:
            # Fallback: top 5 validated competitors ranked on this keyword
            top5 = k.competitors[:5]
            for cid in top5:
                name = name_by_id.get(cid, cid)
                rank = k.competitor_ranks.get(cid)
                comp_lines.append(f"• {name} (#{rank})" if rank else f"• {name}")
        rows.append({
            "Source": k.source,
            "Keyword": k.name,
            "Tier": tier,
            "Score": k.score,
            "Popularity": k.popularity,
            "Search Index": k.total_apps if k.total_apps is not None else "",
            "# Competitors": k.competitor_count,
            "Avg rank": k.avg_competitor_ranking,
            "Bidding apps": k.ad_count if k.ad_count is not None else "",
            "Competitors": "\n".join(comp_lines),
        })
    return pd.DataFrame(rows)


def _competitors_to_df(items: list[Competitor], country: int) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "App ID": c.app_id,
            "Name": c.name,
            "Category": c.category,
            "Seed overlap %": round(c.seed_overlap * 100, 1),
            "Matched seeds": ", ".join(c.matched_seeds),
            "Validated": "YES" if c.validated else "NO",
            "Keywords scraped": len(c.keywords),
            "Upup link": UPUP_APP_URL.format(app_id=c.app_id, country=country),
        }
        for c in items
    ])


def _format_keyword_sheet(sheet, df: pd.DataFrame) -> None:
    if df.empty:
        return
    headers = list(df.columns)
    last_row = sheet.max_row
    if last_row < 2:
        return

    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    widths = {
        "Source": 12,
        "Keyword": 28,
        "Tier": 12,
        "Score": 12,
        "Popularity": 15,
        "Search Index": 14,
        "# Competitors": 17,
        "Avg rank": 13,
        "Bidding apps": 14,
        "Competitors": 52,
    }
    for col_idx, name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = widths.get(name, 14)

    tier_col = headers.index("Tier") + 1 if "Tier" in headers else None
    source_col = headers.index("Source") + 1 if "Source" in headers else None
    score_col = headers.index("Score") + 1 if "Score" in headers else None
    competitors_col = headers.index("Competitors") + 1 if "Competitors" in headers else None
    numeric_cols = {
        headers.index(c) + 1 for c in ("Score", "Popularity", "Search Index", "# Competitors", "Avg rank", "Bidding apps")
        if c in headers
    }

    for row_idx in range(2, last_row + 1):
        zebra = (row_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if col_idx == competitors_col:
                cell.alignment = WRAP_ALIGNMENT
            elif col_idx in numeric_cols or col_idx == tier_col:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = Alignment(vertical="center")
            if col_idx == score_col and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
            if zebra and col_idx != tier_col:
                cell.fill = ZEBRA_FILL

        if tier_col:
            tier_cell = sheet.cell(row=row_idx, column=tier_col)
            v = (tier_cell.value or "").upper()
            if v == "BEST":
                tier_cell.fill = TIER_BEST_FILL
                tier_cell.font = TIER_BEST_FONT
            elif v == "MEDIUM":
                tier_cell.fill = TIER_MED_FILL
                tier_cell.font = TIER_MED_FONT
            else:
                tier_cell.fill = TIER_LOW_FILL
                tier_cell.font = TIER_LOW_FONT

        if source_col:
            src_cell = sheet.cell(row=row_idx, column=source_col)
            if (src_cell.value or "").upper() == "ASO+ASA":
                src_cell.font = FONT_ORANGE

    score_idx = headers.index("Score") + 1 if "Score" in headers else None
    pop_idx = headers.index("Popularity") + 1 if "Popularity" in headers else None
    si_idx = headers.index("Search Index") + 1 if "Search Index" in headers else None
    count_idx = headers.index("# Competitors") + 1 if "# Competitors" in headers else None
    rank_idx = headers.index("Avg rank") + 1 if "Avg rank" in headers else None
    asa_idx = headers.index("Bidding apps") + 1 if "Bidding apps" in headers else None

    for row_idx in range(2, last_row + 1):
        if score_idx:
            cell = sheet.cell(row=row_idx, column=score_idx)
            cell.font = _color_by_thresholds(cell.value, (3.0, 1.0))
        if pop_idx:
            cell = sheet.cell(row=row_idx, column=pop_idx)
            cell.font = _color_by_thresholds(cell.value, (40, 20))
        if si_idx:
            cell = sheet.cell(row=row_idx, column=si_idx)
            cell.font = _color_by_thresholds(cell.value, (3000, 1000))
        if count_idx:
            cell = sheet.cell(row=row_idx, column=count_idx)
            cell.font = _color_by_thresholds(cell.value, (3, 2))
        if rank_idx:
            cell = sheet.cell(row=row_idx, column=rank_idx)
            cell.font = _color_by_thresholds(cell.value, (30, 70), reverse=True)
        if asa_idx:
            cell = sheet.cell(row=row_idx, column=asa_idx)
            cell.font = _color_by_thresholds(cell.value, (3, 8), reverse=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def _hyperlink_competitor_ids(sheet, competitors: list[Competitor], country: int) -> None:
    headers = [cell.value for cell in sheet[1]]
    try:
        app_id_col = headers.index("App ID") + 1
        link_col = headers.index("Upup link") + 1
    except ValueError:
        return

    for row_idx, comp in enumerate(competitors, start=2):
        url = UPUP_APP_URL.format(app_id=comp.app_id, country=country)

        id_cell = sheet.cell(row=row_idx, column=app_id_col)
        id_cell.hyperlink = url
        id_cell.font = LINK_FONT

        link_cell = sheet.cell(row=row_idx, column=link_col)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
