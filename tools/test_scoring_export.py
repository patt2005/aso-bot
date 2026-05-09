"""Self-contained smoke test for scoring + classification + xlsx export.

Builds fake competitors with handcrafted keywords designed to land in each
tier (BEST / MEDIUM / TRASH), runs the pipeline end-to-end, then re-opens
the generated xlsx with openpyxl to confirm it's well-formed.

Run:  python3 tools/test_scoring_export.py
"""
from __future__ import annotations

import sys
from pathlib import Path

# Make sure we can import `core.*` when invoked from the repo root.
REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core.models import Competitor, Keyword
from core.scorer import aggregate_keywords, classify
from core.exporter import to_excel


OUTPUT_PATH = REPO_ROOT / "output" / "test_mock.xlsx"


def _build_fake_competitors() -> list[Competitor]:
    """Construct 5 competitors with overlapping + unique keywords.

    BEST candidates  : shared across >=3 competitors, popularity >= 30, avg rank <= 15
    MEDIUM candidates: shared across exactly 2 competitors, popularity 15-25
    TRASH candidates : unique to one competitor, popularity < 15
    """
    # --- shared "BEST" tier keywords (high pop, top rankings, in 4 competitors) ---
    best_keywords_template = [
        ("fitness tracker", 65.0),
        ("workout planner", 55.0),
        ("calorie counter", 70.0),
    ]

    # --- shared "MEDIUM" tier keywords (in 2 competitors, mid pop) ---
    medium_keywords_template = [
        ("home workout", 22.0),
        ("yoga timer", 18.0),
        ("step counter", 20.0),
    ]

    # --- "TRASH" keywords are competitor-unique, low pop (<15) ---
    trash_per_competitor = {
        0: [("morning stretch", 8.0), ("desk exercise", 6.0)],
        1: [("gym log book", 9.0)],
        2: [("hiking pace", 5.0), ("treadmill timer", 11.0)],
        3: [("dance fitness", 7.0)],
        4: [("rope skipping", 10.0), ("pull up tracker", 12.0)],
    }

    competitors: list[Competitor] = []
    for i in range(5):
        kws: list[Keyword] = []

        # All 5 competitors share the BEST keywords with top-tier rankings (1-10).
        # That gives competitor_count = 5 (>= 3) and avg ranking ~5 (<= 15).
        for name, pop in best_keywords_template:
            kws.append(Keyword(name=name, ranking=2 + i, change=0, popularity=pop))

        # Only competitors 0 and 1 share MEDIUM keywords -> count == 2.
        if i in (0, 1):
            for name, pop in medium_keywords_template:
                kws.append(Keyword(name=name, ranking=20 + i * 3, change=1, popularity=pop))

        # Each competitor has its own unique low-pop trash keywords.
        for name, pop in trash_per_competitor[i]:
            kws.append(Keyword(name=name, ranking=80 + i, change=-1, popularity=pop))

        competitors.append(
            Competitor(
                app_id=f"id-{1000 + i}",
                name=f"FakeFit {i}",
                category="Health & Fitness",
                seed_overlap=0.4 + 0.05 * i,
                matched_seeds=["fitness", "workout"],
                validated=True,
                keywords=kws,
            )
        )

    return competitors


def _print_tier_sample(tier: str, items, n: int = 3) -> None:
    print(f"\n  [{tier}] count={len(items)}")
    for kw in items[:n]:
        print(
            f"    - {kw.name!r:30s} score={kw.score:>7.3f} "
            f"pop={kw.popularity:>5.1f} comps={kw.competitor_count} "
            f"avg_rank={kw.avg_competitor_ranking:>5.1f}"
        )


def _verify_xlsx(path: Path) -> None:
    """Re-open the xlsx with openpyxl to confirm sheets exist and have content."""
    import openpyxl

    assert path.exists(), f"xlsx file was not created at {path}"

    wb = openpyxl.load_workbook(path)
    expected_sheets = {"BEST", "MEDIUM", "COMPETITORS"}
    actual_sheets = set(wb.sheetnames)
    assert expected_sheets == actual_sheets, (
        f"sheet names mismatch. expected={expected_sheets}, actual={actual_sheets}"
    )

    print("\n[XLSX VERIFICATION]")
    for sheet_name in ["BEST", "MEDIUM", "COMPETITORS"]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\n  Sheet {sheet_name!r}: {ws.max_row} rows x {ws.max_column} cols")
        for i, row in enumerate(rows[:3]):
            label = "header" if i == 0 else f"row{i}"
            print(f"    {label}: {row}")

    # BEST sheet must have at least header + 1 data row
    best_ws = wb["BEST"]
    assert best_ws.max_row >= 2, (
        f"BEST sheet is empty (max_row={best_ws.max_row}); expected at least one data row"
    )


def main() -> int:
    competitors = _build_fake_competitors()
    print(f"[SETUP] built {len(competitors)} fake competitors")
    total_kws = sum(len(c.keywords) for c in competitors)
    print(f"[SETUP] total keyword entries across competitors: {total_kws}")

    scored = aggregate_keywords(competitors)
    print(f"[SCORE] aggregated into {len(scored)} unique scored keywords")

    classified = classify(scored)
    best = classified["BEST"]
    medium = classified["MEDIUM"]
    trash = classified["TRASH"]

    print("\n[TIER COUNTS]")
    print(f"  BEST  : {len(best)}")
    print(f"  MEDIUM: {len(medium)}")
    print(f"  TRASH : {len(trash)}")

    # Tier presence assertions
    assert len(best) >= 1, f"expected >= 1 BEST keyword, got {len(best)}"
    assert len(medium) >= 1, f"expected >= 1 MEDIUM keyword, got {len(medium)}"
    assert len(trash) >= 1, f"expected >= 1 TRASH keyword, got {len(trash)}"

    _print_tier_sample("BEST", best)
    _print_tier_sample("MEDIUM", medium)
    _print_tier_sample("TRASH", trash)

    # Export to xlsx
    out = to_excel(classified, competitors, OUTPUT_PATH)
    print(f"\n[EXPORT] wrote {out}")

    _verify_xlsx(out)

    print("\n[OK] all assertions passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
