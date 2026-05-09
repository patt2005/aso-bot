"""Smoke test for the premium Excel exporter.

Builds fake competitors, runs the scoring pipeline, exports a polished xlsx
through `core.exporter_pro.to_excel_pro`, then re-opens it with openpyxl and
checks: expected sheets, title in Summary!A1, frozen pane on BEST, presence
of conditional-formatting rules on BEST's Score column.

Run:  python3 tools/test_exporter_pro.py
"""
from __future__ import annotations

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from core.models import Competitor, Keyword
from core.scorer import aggregate_keywords, classify
from core.exporter_pro import to_excel_pro


OUTPUT_PATH = REPO_ROOT / "output" / "test_pro.xlsx"
TITLE = "Mock Report"
SUBTITLE = "Test run"


def _build_fake_competitors() -> list[Competitor]:
    best_keywords_template = [
        ("fitness tracker", 65.0),
        ("workout planner", 55.0),
        ("calorie counter", 70.0),
        ("running coach", 48.0),
    ]
    medium_keywords_template = [
        ("home workout", 22.0),
        ("yoga timer", 18.0),
        ("step counter", 20.0),
    ]
    trash_per_competitor = {
        0: [("morning stretch", 8.0), ("desk exercise", 6.0)],
        1: [("gym log book", 9.0)],
        2: [("hiking pace", 5.0), ("treadmill timer", 11.0)],
        3: [("dance fitness", 7.0)],
        4: [("rope skipping", 10.0), ("pull up tracker", 12.0)],
    }

    competitors: list[Competitor] = []
    for i in range(5):
        kws: list[Keyword] = []
        for name, pop in best_keywords_template:
            kws.append(Keyword(name=name, ranking=2 + i, change=0, popularity=pop))
        if i in (0, 1, 2):
            for name, pop in medium_keywords_template:
                kws.append(Keyword(name=name, ranking=20 + i * 3, change=1, popularity=pop))
        for name, pop in trash_per_competitor[i]:
            kws.append(Keyword(name=name, ranking=80 + i, change=-1, popularity=pop))

        competitors.append(
            Competitor(
                app_id=f"id-{1000 + i}",
                name=f"FakeFit {i}",
                category="Health & Fitness",
                seed_overlap=0.4 + 0.05 * i,
                matched_seeds=["fitness", "workout"],
                validated=True,
                keywords=kws,
            )
        )
    return competitors


def _verify_xlsx(path: Path) -> None:
    import openpyxl

    assert path.exists(), f"xlsx file was not created at {path}"

    wb = openpyxl.load_workbook(path)
    actual_sheets = set(wb.sheetnames)
    required_sheets = {"Summary", "BEST", "MEDIUM", "COMPETITORS", "Charts"}
    missing = required_sheets - actual_sheets
    assert not missing, f"missing sheets: {missing} (actual={actual_sheets})"

    print("\n[XLSX VERIFICATION]")
    print(f"  sheets: {wb.sheetnames}")

    summary_ws = wb["Summary"]
    a1 = summary_ws["A1"].value
    assert a1 is not None and TITLE in str(a1), (
        f"Summary!A1 should contain title {TITLE!r}, got {a1!r}"
    )
    print(f"  Summary!A1: {a1!r}")

    best_ws = wb["BEST"]
    assert best_ws.freeze_panes is not None, "BEST sheet should have frozen panes"
    assert str(best_ws.freeze_panes).startswith("A2"), (
        f"BEST sheet freeze_panes should start at A2, got {best_ws.freeze_panes!r}"
    )
    print(f"  BEST.freeze_panes: {best_ws.freeze_panes}")

    cf_rules = best_ws.conditional_formatting
    rule_ranges = list(cf_rules._cf_rules.keys()) if hasattr(cf_rules, "_cf_rules") else list(cf_rules)
    assert len(rule_ranges) > 0, "BEST sheet should have conditional formatting rules"

    score_col_has_rule = False
    for rng in rule_ranges:
        rng_str = str(rng) if hasattr(rng, "__str__") else rng
        if rng_str.startswith("B2:B") or "B2:B" in rng_str:
            score_col_has_rule = True
            break
    assert score_col_has_rule, (
        f"BEST sheet should have conditional formatting on the Score column (B), got ranges: {rule_ranges}"
    )
    print(f"  BEST conditional-formatting rule ranges: {[str(r) for r in rule_ranges]}")

    comp_ws = wb["COMPETITORS"]
    assert comp_ws.max_row >= 2, "COMPETITORS sheet should have at least one row"
    a2 = comp_ws["A2"]
    assert a2.hyperlink is not None, "COMPETITORS App ID column should have hyperlinks"
    print(f"  COMPETITORS!A2 hyperlink target: {a2.hyperlink.target}")


def main() -> int:
    competitors = _build_fake_competitors()
    print(f"[SETUP] built {len(competitors)} fake competitors")

    scored = aggregate_keywords(competitors)
    print(f"[SCORE] aggregated into {len(scored)} unique scored keywords")

    classified = classify(scored)
    for tier in ("BEST", "MEDIUM", "TRASH"):
        print(f"  {tier}: {len(classified.get(tier, []))}")

    out = to_excel_pro(
        classified,
        competitors,
        OUTPUT_PATH,
        title=TITLE,
        subtitle=SUBTITLE,
    )
    print(f"\n[EXPORT] wrote {out}")

    _verify_xlsx(out)
    print("\nOK")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
